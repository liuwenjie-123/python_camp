"""
1、初始化表格信息
2、按项目去ping，
 2.1 连续5次ping不通的计为down
3、根据不通的生成信息
 3.1 记录离线率到offline.xlsx
4、发送钉钉消息

"""

import time, datetime
import prettytable as pt
from openpyxl import load_workbook
from multiping import multi_ping
from dingtalkchatbot.chatbot import DingtalkChatbot
from concurrent.futures import ThreadPoolExecutor




class vm_fping:
    def __init__(self):
        # excel表格路径
        self.wb = load_workbook(r'vm01.xlsx')
        # 离线率信息
        self.wb_off = load_workbook(r'offline.xlsx')
        # {"电讯云"={"jh"=[{"host_ip"：host_ip, "placement" : placement, "vm_ip" : vm_ip, "down" : fales},"xx"},  }
        self.vm_dict = {}
        # {"项目_.区域"：[虚拟机ip列表]}
        self.vm_data = {}
        # down_info  {"项目":{"区域": [["机位", "物理机ip", "虚拟机ip"], [] ] } }
        self.down_info = {}
        # self.offline 离线率信息  {"项目":[["区域", "离线数", "总数", "离线率"], ["gy", "1", "100", "1%"] ] ] }
        self.offline = {}
        # self.ping_count  ping次数
        self.ping_count = 5
        # 每次ping间隔，单位s
        self.ping_sleep = 60

    def get_ip_dict(self):
        # 遍历excel构造self.vm_dict字典， {"电讯云"={"jh"=[{"host_ip"：host_ip, "placement" : placement, "vm_ip" : vm_ip, "down" : fales},"xx"}, }
        for sheet in self.wb:
            for i, row in enumerate(sheet.rows):
                if i == 0:
                    continue

                self.vminfo = {"host_ip": row[2].value, "placement": row[3].value, "vm_ip": row[4].value, "down" : False}
                # 判断项目是否存在
                if not self.vm_dict.get(row[0].value):
                    # 初始化 {项目 ： {} } 结构
                    self.vm_dict[row[0].value] = {}

                # 判断项目-区域是否存在
                if not self.vm_dict.get(row[0].value).get(row[1].value):
                    # 初始化 {项目 ： {jh:[]} } 结构
                    self.vm_dict[row[0].value][row[1].value] = []

                # 添加数据
                self.vm_dict[row[0].value][row[1].value].append(self.vminfo)

                # print(row[0].value)
                # print(row[1].value)
                # print(row[2].value)
                # print(row[3].value)
                # print(row[4].value)

        # print(self.vm_dict)


    def get_vm_data(self):
        # 构造self.vm_data字典，{"项目_.区域"：[虚拟机ip列表]}
        for projcte, k in self.vm_dict.items():
            for zone, vm_list_tmp in k.items():
                for vm in vm_list_tmp:
                    if not self.vm_data.get(projcte + "_." + zone):
                        self.vm_data[projcte + "_." + zone] = []
                    self.vm_data[projcte + "_." + zone].append(vm["vm_ip"])
        # self.ip_list = []
        # mp = MultiPing(self.ip_list)
        # mp.send()
        # responses, no_responses = mp.receive(1)
        # print(self.vm_data)


    def pre_exec_fping(self):
        # 预执行ping操作
        pool = ThreadPoolExecutor(10)
        for project_zone, ip_list in self.vm_data.items():
            project = project_zone.split('_.')[0]
            zone = project_zone.split('_.')[1]

            # 构造字典数据down_info,offline
            if not self.down_info.get(project):
                self.down_info[project] = {}
            if not self.down_info[project].get(zone):
                self.down_info[project][zone] = []

            if not self.offline.get(project):
                self.offline[project] = []
            # print("开始执行多线程ping，目前项目区域：{}{}".format(project, zone))
            pool = ThreadPoolExecutor(10)
            pool.submit(self.multi_exec_fping, project, zone, ip_list)
        pool.shutdown()
        # multi_ping有多线程操作，睡眠等待数据执行完成
        time.sleep(90)
        # print("{} 完成ping检查".format(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))))

    def multi_exec_fping(self, project, zone, ip_list):
        # 执行多线程ping操作，并填充self.down_info、self.offline
        # print("project:{}\tzone:{}\tip_list:{}\t".format(project, zone, ip_list))
        names = self.__dict__
        names["ip_list{}{}".format(project, zone)]
        if names["ip_list{}{}".format(project, zone)]:
            # print("ip_list:",ip_list)
            names["no_responses{}{}".format(project, zone)] = []
            for i in range(0, self.ping_count):
                names["responses{}{}".format(project, zone)], names["no_responses_tmp{}{}".format(project, zone)] = multi_ping(names["ip_list{}{}".format(project, zone)], timeout=2, retry=1)
                if i != 0 :
                    names["responses{}{}".format(project, zone)] = list(set(names["no_responses{}{}".format(project, zone)]) & set(names["no_responses_tmp{}{}".format(project, zone)]))
                else:
                    names["no_responses{}{}".format(project, zone)] = names["no_responses_tmp{}{}".format(project, zone)]
                time.sleep(self.ping_sleep)
            # print("responses:",responses)
            # print("no_responses:",no_responses)

        # 填充数据
        # down_info  {"项目":{"区域": [["机位", "物理机ip", "虚拟机ip"], [] ] } }
        # self.offline 离线率信息  {"项目":[["区域", "离线数", "总数", "离线率"], ["gy", "1", "100", "1%"] ] ] }
        for i in names["no_responses{}{}".format(project, zone)]:
            for item in self.vm_dict[project][zone]:
                if i == item["vm_ip"]:
                    self.down_info[project][zone].append([item["placement"], item["host_ip"], item["vm_ip"] ])
        # print("down_info:", self.down_info)

        # 构造offline字典数据
        self.offline[project].append([zone, len(names["no_responses{}{}".format(project, zone)]), len(ip_list), "{:.2%}".format(len(names["no_responses{}{}".format(project, zone)]) / len(ip_list), )])
        # print("{} mu offline:".format(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))), self.offline)



    # 格式化并发送详细离线信息到钉钉
    def sed_down_info(self):
        down_info_total = ""
        for project, project_dict in self.down_info.items():
            down_info_total += "{}\n".format(project.center(40, "#"),)
            down_info_total += "{}\t{}\t{}\n".format("机位", "物理机ip", "虚拟机ip")
            for zoen, info  in project_dict.items():
                for i in info:
                    if not i:
                        continue
                    down_info_total += "{}\t{}\t{}\n".format(i[0], i[1], i[2])
            # down_info_total += "{}\n".format("".center(40, "#"),)
        # print("down_info_total:",down_info_total)
        # self.sed_dd_msg(down_info_total)


    # 格式化，并发送offline信息到钉钉
    def sed_info_to_dd(self):
        for project_name, zone_info in self.offline.items():
            # 写数据到exce表
            sheet1 = self.wb_off.worksheets[0]
            for v in zone_info:
                values = [datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), project_name, ]
                values += v
                # print(values)
                sheet1.append(values)

            # 把数据转换成表格
            tb = pt.PrettyTable()
            tb.field_names = ["区域", "离线数", "总数", "离线率"]
            for i in zone_info: tb.add_row(i)
            content = "瑞云-{}\n{}".format(project_name, tb)

            # print("content:", content)
            # self.sed_dd_msg(content)


    # 钉钉发送消息的函数
    def sed_dd_msg(self, content ):
        # WebHook地址
        webhook = 'https://oapi.dingtalk.com/robot/send?access_token=8240c507ed559969a035ff87bd9ca7c0e6f9cee3857a5e8e5be9a305f09dfe65'
        # 初始化机器人小丁
        xiaoding = DingtalkChatbot(webhook, secret='SEC876cf49b4b17a827b86fea8b35a4a3e0706bba13f576185ace0680030613ae34')
        # Text消息@所有人
        xiaoding.send_text(msg=content, is_at_all=False)
        # xiaoding.send_link(title='万万没想到，李小璐竟然...', text='故事是这样子的...',
        #                    message_url='http://www.kwongwah.com.my/?p=454748", pic_url="https://pbs.twimg.com/media/CEwj7EDWgAE5eIF.jpg')
        # xiaoding.send_markdown(title='氧气文字', text=content)






if __name__ == '__main__':
    vm_cache = vm_fping()
    vm_cache.get_ip_dict()
    vm_cache.get_vm_data()
    vm_cache.pre_exec_fping()
    vm_cache.sed_down_info()
    vm_cache.sed_info_to_dd()
    vm_cache.wb_off.save(r"offline.xlsx")
    print("完成")












